from __future__ import annotations

import hashlib
import queue
import re
import ssl
import threading
import time
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, timedelta
from email.utils import parsedate_to_datetime
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus, urljoin

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from utils import clean_article_url, normalize_text, normalize_url, parse_date


DEFAULT_HEADERS = [
    "编号", "标题", "来源", "原文链接", "发布时间", "收集时间", "语言", "地区", "类别",
    "中文摘要", "关键实体", "行业相关性", "重要性评分", "内容角度", "推荐用途",
    "状态", "是否需人工核验", "是否今日推送", "是否已推送", "推送时间", "人工备注",
]


FIELD_TO_HEADER = {
    "id": "编号",
    "title": "标题",
    "source": "来源",
    "url": "原文链接",
    "published_date": "发布时间",
    "collected_date": "收集时间",
    "language": "语言",
    "region": "地区",
    "category": "类别",
    "summary_cn": "中文摘要",
    "key_entities": "关键实体",
    "relevance_to_zettalab": "行业相关性",
    "importance_score": "重要性评分",
    "content_angle": "内容角度",
    "platform": "推荐用途",
    "status": "状态",
    "need_verify": "是否需人工核验",
    "push_today": "是否今日推送",
    "pushed": "是否已推送",
    "pushed_at": "推送时间",
    "reviewer_note": "人工备注",
}

HEADER_TO_FIELD = {value: key for key, value in FIELD_TO_HEADER.items()}
HEADER_TO_FIELD.update({
    "与 ZettaLab 相关性": "relevance_to_zettalab",
    "与ZettaLab相关性": "relevance_to_zettalab",
    "相关性": "relevance_to_zettalab",
    "推荐平台": "platform",
    "平台": "platform",
})


@dataclass
class FetchResult:
    fetched: int
    appended: int
    skipped_existing: int
    skipped_filtered: int
    errors: list[str]


@dataclass
class RawNewsItem:
    title: str
    link: str
    published: date | None = None
    summary: str = ""


class _NewsLinkParser(HTMLParser):
    """从公开新闻稿列表页提取链接文字，避免引入浏览器自动化。"""

    def __init__(self) -> None:
        super().__init__()
        self._href = ""
        self._chunks: list[str] = []
        self.links: list[tuple[str, str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        attrs_map = {key.lower(): value or "" for key, value in attrs}
        self._href = attrs_map.get("href", "")
        self._chunks = []

    def handle_data(self, data: str) -> None:
        if self._href:
            self._chunks.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() != "a" or not self._href:
            return
        title = normalize_text(unescape(" ".join(self._chunks)))
        if title:
            self.links.append((self._href, title))
        self._href = ""
        self._chunks = []


def _strip_html(text: str) -> str:
    text = re.sub(r"<[^>]+>", " ", text or "")
    return normalize_text(unescape(text))


def _parse_feed_date(value: str) -> date | None:
    value = normalize_text(value)
    if not value:
        return None
    try:
        return parsedate_to_datetime(value).date()
    except Exception:
        return parse_date(value)


def _text(node: ET.Element, tag_names: list[str]) -> str:
    wanted = set(tag_names)
    for child in list(node):
        local = child.tag.split("}", 1)[-1].lower()
        if local in wanted:
            return normalize_text(child.text)
    return ""


def _link(node: ET.Element) -> str:
    for child in list(node):
        local = child.tag.split("}", 1)[-1].lower()
        if local != "link":
            continue
        href = child.attrib.get("href")
        if href:
            return normalize_text(href)
        if child.text:
            return normalize_text(child.text)
    return ""


def _download_xml(url: str, timeout: int = 12, retries: int = 2) -> ET.Element:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "ZettaLab-Biomed-Radar/1.0 (+local RSS reader)",
            "Accept": "application/rss+xml, application/atom+xml, application/xml, text/xml;q=0.9, */*;q=0.8",
        },
    )
    try:
        import certifi

        context = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        context = ssl.create_default_context()
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(request, timeout=timeout, context=context) as response:
                data = response.read()
            break
        except Exception as exc:
            last_error = exc
            if attempt >= retries:
                raise
            time.sleep(1.5 * (attempt + 1))
    else:
        raise last_error or RuntimeError("RSS download failed")
    try:
        return ET.fromstring(data)
    except ET.ParseError as exc:
        prefix = data[:80].decode("utf-8", errors="ignore").replace("\n", " ")
        raise ValueError(f"RSS/Atom XML 解析失败：{exc}；返回内容开头：{prefix}") from exc


def _download_text(url: str, timeout: int = 10) -> str:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (compatible; BiomedRadar/1.0; +local news monitor)",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        },
    )
    try:
        import certifi

        context = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        context = ssl.create_default_context()
    with urllib.request.urlopen(request, timeout=timeout, context=context) as response:
        data = response.read()
        charset = response.headers.get_content_charset() or "utf-8"
    return data.decode(charset, errors="ignore")


def _extract_items(root: ET.Element) -> list[ET.Element]:
    items = []
    for node in root.iter():
        local = node.tag.split("}", 1)[-1].lower()
        if local in {"item", "entry"}:
            items.append(node)
    return items


def _extract_feed_items(root: ET.Element, max_items: int) -> list[RawNewsItem]:
    rows: list[RawNewsItem] = []
    for item in _extract_items(root)[:max_items]:
        title = _text(item, ["title"])
        link = _link(item)
        published_raw = _text(item, ["pubdate", "published", "updated", "date"])
        summary = _strip_html(_text(item, ["description", "summary", "content", "encoded"]))
        rows.append(RawNewsItem(title=title, link=link, published=_parse_feed_date(published_raw), summary=summary))
    return rows


def _parse_date_near_html_link(html: str, link: str, title: str) -> date | None:
    """从链接附近的 HTML 片段中猜测发布日期；失败时返回 None，交给人工核验。"""
    idx = html.find(link)
    if idx < 0 and title:
        idx = html.find(title)
    if idx < 0:
        return None
    snippet = _strip_html(html[max(0, idx - 800): idx + 1200])
    date_patterns = [
        r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},\s+\d{4}\b",
        r"\b\d{4}-\d{1,2}-\d{1,2}\b",
        r"\b\d{1,2}/\d{1,2}/\d{4}\b",
        r"\b\d{4}年\d{1,2}月\d{1,2}日\b",
    ]
    for pattern in date_patterns:
        match = re.search(pattern, snippet, re.IGNORECASE)
        if match:
            parsed = parse_date(match.group(0))
            if parsed:
                return parsed
    return None


def _extract_html_list_items(url: str, max_items: int, source: dict[str, Any] | None = None) -> list[RawNewsItem]:
    html = _download_text(url)
    parser = _NewsLinkParser()
    parser.feed(html)
    rows: list[RawNewsItem] = []
    seen: set[str] = set()
    source = source or {}
    release_markers = source.get("link_include_patterns") or [
        "/news-releases/",
        "/news-release/",
        "/press-release/",
        "/en/news-release/",
        "/newsroom/",
    ]
    link_exclude_patterns = source.get("link_exclude_patterns", [])
    title_include_keywords = source.get("title_include_keywords", [])
    skip_titles = {
        "read more", "next page", "view all", "contact us", "sign up", "rss", "news",
        "browse news releases", "all news releases",
    }
    for href, title in parser.links:
        absolute = urljoin(url, href)
        normalized_url = normalize_url(absolute)
        normalized_title = normalize_text(title)
        if not normalized_url or normalized_url in seen:
            continue
        if len(normalized_title) < 18 or normalized_title.lower() in skip_titles:
            continue
        if not any(marker in normalized_url.lower() for marker in release_markers):
            continue
        if any(marker.lower() in normalized_url.lower() for marker in link_exclude_patterns):
            continue
        if title_include_keywords and not any(_term_hit(normalized_title, keyword) for keyword in title_include_keywords):
            continue
        seen.add(normalized_url)
        rows.append(RawNewsItem(
            title=normalized_title,
            link=normalized_url,
            published=_parse_date_near_html_link(html, href, normalized_title),
            summary="公开新闻稿列表页自动检索；需打开原文确认正文细节。",
        ))
        if len(rows) >= max_items:
            break
    return rows


def _term_hit(text: str, term: str) -> bool:
    text_lower = text.lower()
    term_lower = term.lower().strip()
    if not term_lower:
        return False
    if re.search(r"[a-z0-9]", term_lower):
        pattern = r"(?<![a-z0-9])" + re.escape(term_lower) + r"(?![a-z0-9])"
        return re.search(pattern, text_lower) is not None
    return term_lower in text_lower


def _keyword_hit(text: str, include_keywords: list[str], exclude_keywords: list[str]) -> bool:
    if exclude_keywords and any(_term_hit(text, k) for k in exclude_keywords):
        return False
    if not include_keywords:
        return True
    return any(_term_hit(text, k) for k in include_keywords)


def _fallback_search_url(title: str) -> str:
    return "https://www.baidu.com/s?wd=" + quote_plus(title)


def _resolve_google_news_url(link: str, title: str, timeout: int = 8) -> tuple[str, bool]:
    """把 Google News 聚合链接尽量还原为原文链接；失败时返回国内搜索入口。"""
    if "news.google.com" not in normalize_text(link).lower():
        return link, False
    result_queue: queue.Queue[tuple[str, bool]] = queue.Queue(maxsize=1)

    def worker() -> None:
        try:
            from googlenewsdecoder import gnewsdecoder

            result = gnewsdecoder(link)
            decoded_url = normalize_url(result.get("decoded_url")) if isinstance(result, dict) and result.get("status") else ""
            ok = bool(decoded_url and "news.google.com" not in decoded_url.lower())
            result_queue.put((decoded_url, ok), block=False)
        except Exception:
            result_queue.put(("", False), block=False)

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()
    try:
        decoded_url, ok = result_queue.get(timeout=timeout)
        if ok:
            return decoded_url, True
    except queue.Empty:
        pass
    return _fallback_search_url(title), False


def _score_and_angle(
    title: str,
    summary: str,
    category: str,
    source_name: str,
    source_kind: str,
    published: date | None,
    run_date: date,
    config: dict[str, Any],
) -> tuple[int, str, str, str, str, str, bool]:
    text = f"{title} {summary} {category}".lower()
    fetch_cfg = config.get("fetch", {})
    preferred_days = int(fetch_cfg.get("preferred_days", 3))

    verify_terms = [
        "fda", "ema", "nmpa", "cde", "approval", "clinical", "trial", "phase",
        "financing", "funding", "regulatory", "license-out", "license-in",
        "监管", "审批", "审评", "临床", "融资", "授权", "首付款", "里程碑", "合作金额",
    ]
    high_terms = [
        "drug discovery", "clinical trial", "gene editing", "cell therapy", "protein",
        "创新药", "获批", "上市申请", "临床试验", "授权", "出海", "bd", "license-out",
        "adc", "双抗", "多抗", "核酸药物", "sirna", "mrna", "car-t", "基因编辑",
        "细胞治疗", "ai制药", "人工智能制药", "医保谈判", "集采", "商业化",
    ]
    industry_terms = [
        "创新药", "获批", "上市申请", "优先审评", "突破性疗法", "临床试验",
        "ii期", "iii期", "2期", "3期", "关键临床", "顶线数据",
        "license-out", "license-in", "授权", "出海", "首付款", "里程碑", "总交易金额",
        "adc", "双抗", "多抗", "核酸药物", "sirna", "mrna", "car-t", "tcr-t",
        "基因编辑", "细胞治疗", "医保谈判", "集采", "商业化", "cde", "nmpa",
    ]
    event_terms = [
        "获批", "批准", "受理", "纳入优先审评", "优先审评", "上市申请", "ind", "nda", "bla",
        "达到主要终点", "主要终点", "顶线数据", "公布数据", "临床数据", "iii期", "3期",
        "ii期", "2期", "完成入组", "启动临床", "获准临床", "签署", "达成合作", "授权",
        "license-out", "license-in", "首付款", "里程碑", "交易金额", "完成融资", "融资",
        "并购", "收购", "出售", "剥离", "管线调整", "终止", "暂停", "裁员",
        "集采中选", "医保谈判", "纳入医保", "商业化", "获受理",
    ]
    non_event_terms = [
        "报告", "白皮书", "趋势", "现状", "格局", "怎么看", "解读", "策略", "实践",
        "回顾", "沙龙", "论坛", "会议", "齐聚", "活动", "路演", "指数", "股票",
    ]

    priority_terms = fetch_cfg.get("priority_keywords", [])
    low_value_terms = fetch_cfg.get("low_value_keywords", [])
    need_verify = any(_term_hit(text, term) for term in verify_terms)
    relevance = "高" if any(_term_hit(text, term) for term in high_terms) else "中"
    priority_hits = sum(1 for term in priority_terms if _term_hit(text, term))
    low_value_hits = sum(1 for term in low_value_terms if _term_hit(text, term))
    industry_hits = sum(1 for term in industry_terms if _term_hit(text, term))
    event_hits = sum(1 for term in event_terms if _term_hit(text, term))
    non_event_hits = sum(1 for term in non_event_terms if _term_hit(text, term))

    score = 3
    if relevance == "高" or need_verify:
        score = 4
    if priority_hits >= 2 or industry_hits >= 2:
        score = 5
    elif priority_hits == 1 or industry_hits == 1:
        score = max(score, 4)

    source_lower = source_name.lower()
    source_kind = normalize_text(source_kind).lower()
    if any(name in source_lower for name in ["fda", "ema", "nih"]):
        score = min(score + 1, 5)
    if source_kind == "primary" and not any(name in source_lower for name in ["nih"]):
        score = min(score + 1, 5)
    if "google news" in source_lower:
        score = max(score - 1, 2)
    if low_value_hits:
        score = max(score - min(low_value_hits, 2), 2)
    if "google news" in source_lower and low_value_hits:
        score = min(score, 3)
    if config.get("editorial_profile", {}).get("event_required_for_top", True):
        if event_hits == 0 and "google news" in source_lower:
            score = min(score, 3)
        if event_hits == 0 and source_kind == "aggregator":
            score = min(score, 3)
        if non_event_hits and event_hits == 0:
            score = min(score, 3)
        if non_event_hits >= 2:
            score = min(score, 3)

    if published:
        age_days = (run_date - published).days
        if age_days > preferred_days and priority_hits < 2 and industry_hits < 2:
            score = min(score, 3)
        if age_days < 0:
            score = min(score, 2)

    status = "建议推送" if score >= 4 and low_value_hits == 0 else "已归档"

    if any(_term_hit(text, term) for term in ["license-out", "license-in", "授权", "出海", "首付款", "里程碑"]):
        angle = "药企 BD / 授权出海动态，可观察创新药全球化和资产价值兑现。"
    elif any(_term_hit(text, term) for term in ["获批", "上市申请", "优先审评", "cde", "nmpa", "药审"]):
        angle = "创新药审评审批动态，需关注适应症、审评阶段和正式公告口径。"
    elif any(_term_hit(text, term) for term in ["临床试验", "iii期", "3期", "顶线数据", "clinical trial"]):
        angle = "临床数据或关键试验进展，需核验终点、人群和统计口径。"
    elif any(_term_hit(text, term) for term in ["adc", "双抗", "多抗", "核酸药物", "sirna", "mrna"]):
        angle = "前沿药物平台进展，可观察技术路线、适应症布局和管线竞争。"
    elif any(_term_hit(text, term) for term in ["ai", "artificial intelligence", "machine learning", "ai制药", "人工智能制药"]):
        angle = "AI 制药或数据驱动研发进展，可观察其从概念验证走向流程落地的程度。"
    elif any(_term_hit(text, term) for term in ["crispr", "gene editing", "基因编辑", "细胞治疗", "car-t"]):
        angle = "细胞与基因治疗进展，需关注适应症、临床阶段和可及性。"
    elif any(_term_hit(text, term) for term in ["医保谈判", "集采", "商业化", "准入"]):
        angle = "药企商业化、医保准入与支付政策动态，可观察产品放量环境。"
    else:
        angle = f"{category} 相关动态，可作为行业观察素材。"

    if "google news" in source_lower:
        note = "Google News 聚合线索；链接已改为国内搜索入口，需从搜索结果进入最终原始来源核验。"
    elif source_kind == "primary":
        note = "一级公开来源自动检索；适合优先阅读原文并核验关键数字。"
    elif need_verify:
        note = "涉及审批、临床、融资、授权或监管信息；需以原文和权威来源核验。"
    else:
        note = "RSS 自动检索生成；建议人工阅读原文后再外部引用。"

    selected_for_table = bool(score >= 4 or need_verify or fetch_cfg.get("include_archived", True))
    return score, relevance, "是" if need_verify else "否", angle, note, status, selected_for_table


def fetch_rss_items(config: dict[str, Any], run_date: date) -> tuple[list[dict[str, Any]], list[str], int]:
    fetch_cfg = config.get("fetch", {})
    days_back = int(fetch_cfg.get("days_back", 7))
    since = run_date - timedelta(days=days_back)
    require_direct_article_link = bool(fetch_cfg.get("require_direct_article_link", True))
    include_keywords = fetch_cfg.get("include_keywords", [])
    exclude_keywords = fetch_cfg.get("exclude_keywords", [])
    max_items = int(fetch_cfg.get("max_items_per_source", 20))
    rows: list[dict[str, Any]] = []
    errors: list[str] = []
    fetched_count = 0

    for source in fetch_cfg.get("sources", []):
        if source.get("enabled") is False:
            continue
        source_name = source.get("name", "RSS")
        url = source.get("url", "")
        if not url:
            continue
        try:
            source_type = normalize_text(source.get("type") or source.get("format") or "rss").lower()
            if source_type == "html_list":
                items = _extract_html_list_items(url, max_items, source)
            else:
                root = _download_xml(
                    url,
                    timeout=int(fetch_cfg.get("request_timeout_seconds", 12)),
                    retries=int(fetch_cfg.get("request_retries", 2)),
                )
                items = _extract_feed_items(root, max_items)
        except Exception as exc:
            errors.append(f"{source_name} 拉取失败：{exc}")
            continue

        for item in items:
            fetched_count += 1
            title = item.title
            link = item.link
            summary = item.summary
            published = item.published
            decoded_from_google = False
            if published and published < since:
                continue
            text_for_filter = f"{title} {summary}"
            if not title or not link or not _keyword_hit(text_for_filter, include_keywords, exclude_keywords):
                continue

            score, relevance, need_verify, angle, note, status, selected_for_table = _score_and_angle(
                title,
                summary,
                source.get("category", ""),
                source_name,
                source.get("source_kind", ""),
                published,
                run_date,
                config,
            )
            if not selected_for_table:
                continue
            if "google news" in source_name.lower():
                if score >= 4 or need_verify == "是":
                    link, decoded_from_google = _resolve_google_news_url(
                        link,
                        title,
                        timeout=int(fetch_cfg.get("google_decode_timeout_seconds", 8)),
                    )
                    if not decoded_from_google:
                        link, decoded_from_google = _resolve_google_news_url(
                            item.link,
                            title,
                            timeout=int(fetch_cfg.get("google_decode_timeout_seconds", 8)),
                        )
                else:
                    link = _fallback_search_url(title)
                if decoded_from_google:
                    note = "Google News 聚合线索已解析为原文链接；仍需打开原文核验关键事实。"
                elif require_direct_article_link:
                    score = min(score, 3)
                    status = "已归档"
                    note = "Google News 聚合线索未能解析为原文链接；为避免邮件链接停留在搜索页，本条仅归档不推送。"
                    if not fetch_cfg.get("include_archived", True):
                        continue

            stable_id = hashlib.sha1(normalize_url(link).encode("utf-8")).hexdigest()[:10]
            rows.append({
                "id": stable_id,
                "title": title,
                "source": source_name,
                "url": clean_article_url(link),
                "published_date": published.strftime("%Y-%m-%d") if published else "",
                "collected_date": run_date.strftime("%Y-%m-%d"),
                "language": source.get("language", ""),
                "region": source.get("region", ""),
                "category": source.get("category", ""),
                "summary_cn": summary[:260] if summary else "原始来源未在列表页提供摘要，需人工阅读原文补充。",
                "key_entities": "",
                "relevance_to_zettalab": relevance,
                "importance_score": score,
                "content_angle": angle,
                "platform": "内部简报",
                "status": status,
                "need_verify": need_verify,
                "push_today": "是" if score >= 4 else "否",
                "pushed": "否",
                "pushed_at": "",
                "reviewer_note": note,
            })
    return rows, errors, fetched_count


def ensure_news_workbook(excel_path: Path, sheet_name: str) -> None:
    if excel_path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(DEFAULT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for name in ["来源清单", "分类字典", "每日简报", "使用SOP"]:
        wb.create_sheet(name)
    wb.save(excel_path)


def append_fetched_rows(excel_path: Path, sheet_name: str, rows: list[dict[str, Any]]) -> tuple[int, int]:
    if not rows:
        return 0, 0
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表不存在：{sheet_name}")
    ws = wb[sheet_name]
    headers = [normalize_text(cell.value) for cell in ws[1]]
    existing_urls = set()
    url_col = headers.index("原文链接") + 1 if "原文链接" in headers else None
    if url_col:
        for row_idx in range(2, ws.max_row + 1):
            url = normalize_url(ws.cell(row=row_idx, column=url_col).value)
            if url:
                existing_urls.add(url)

    appended = 0
    skipped = 0
    next_id = ws.max_row
    for row in rows:
        url = normalize_url(row.get("url"))
        if url and url in existing_urls:
            skipped += 1
            continue
        values = []
        next_id += 1
        for header in headers:
            field = HEADER_TO_FIELD.get(header, "")
            value = row.get(field, "")
            if field == "id" and not value:
                value = next_id - 1
            values.append(value)
        ws.append(values)
        if url:
            existing_urls.add(url)
        appended += 1
    wb.save(excel_path)
    return appended, skipped
