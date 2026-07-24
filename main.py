from __future__ import annotations

import argparse
import csv
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from dotenv import load_dotenv

from digest_builder import _feedback_id, build_html, build_plain_text, save_digest_outputs
from excel_loader import backup_excel, load_news_excel, write_push_status
from feedback_loader import audit_feedback_history, load_feedback_weights
from feedback_mail_collector import collect_feedback_mail
from feedback_portal import build_feedback_portal
from feedback_cloud import publish_digest, sync_feedback
from logging_config import setup_logging
from mailer import send_mail
from rss_fetcher import append_fetched_rows, ensure_news_workbook, fetch_rss_items
from utils import clean_article_url, ensure_dirs, load_config, parse_run_date, resolve_path
from validator import select_digest_items, validate_rows


BASE_DIR = Path(__file__).resolve().parent


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="ZettaLab 生物医药资讯雷达")
    parser.add_argument("--excel", required=True, help="资讯库 Excel 文件路径")
    parser.add_argument("--config", default=str(BASE_DIR / "config.yaml"), help="配置文件路径")
    parser.add_argument("--date", default=None, help="指定日报日期，格式 YYYY-MM-DD")
    parser.add_argument("--dry-run", action="store_true", help="只生成日报，不发送邮件、不写回 Excel")
    parser.add_argument("--send", action="store_true", help="发送邮件；发送成功后写回 Excel")
    parser.add_argument("--validate-only", action="store_true", help="仅检查 Excel 数据质量")
    parser.add_argument("--fetch", action="store_true", help="先从 config.yaml 配置的公开 RSS 源自动检索资讯并写入 Excel")
    parser.add_argument("--fetch-only", action="store_true", help="只自动检索并写入 Excel，不生成日报")
    parser.add_argument("--verbose", action="store_true", help="输出调试信息")
    return parser.parse_args()


def _log_issues(logger, issues) -> None:
    """把字段质量问题写入日志。"""
    if not issues:
        logger.info("字段质量检查：未发现明显问题。")
        return
    for issue in issues:
        logger.warning("第 %s 行 [%s] %s：%s", issue.row, issue.level, issue.field, issue.message)


def _dedupe_pushed_rows(selection) -> list[dict]:
    """合并日报里的重点资讯和核验资讯，避免附件/发送记录重复出现同一条。"""
    rows = []
    seen = set()
    for row in selection.重点资讯 + selection.需人工核验:
        key = row.get("_excel_row") or row.get("url") or f"{row.get('title', '')}|{row.get('source', '')}"
        if key in seen:
            continue
        seen.add(key)
        rows.append(row)
    return rows


def _attachment_verify_reason(row: dict) -> str:
    """给附件里的核验列提供一句运营可读的原因。"""
    if row.get("need_verify") != "是":
        return ""
    note = str(row.get("reviewer_note") or "").strip()
    if note:
        return note
    text = f"{row.get('title', '')} {row.get('summary_cn', '')} {row.get('content_angle', '')}"
    if any(term in text for term in ["NMPA", "CDE", "审批", "审评", "获批", "受理"]):
        return "涉及监管审批或审评进展，需以官方公告或原文核验。"
    if any(term in text for term in ["融资", "并购", "首付款", "里程碑", "交易金额", "授权"]):
        return "涉及融资、并购、授权或交易金额，需核验金额、口径和交易条款。"
    if any(term in text for term in ["临床", "III期", "II期", "数据", "终点"]):
        return "涉及临床数据或试验进展，需核验终点、人群和数据来源。"
    return "涉及高风险事实信息，建议人工打开原文核验后再外部引用。"


def _save_pushed_items_excel(output_dir: Path, run_date, selection) -> Path:
    """把当天进入邮件的资讯另存为 Excel，方便收件人下载和二次筛选。"""
    path = output_dir / f"pushed_items_{run_date:%Y%m%d}.xlsx"
    headers = [
        "反馈编号", "内容价值", "原因标签", "补充备注", "运行日期", "Excel行号", "标题", "来源", "原文链接", "发布时间", "地区", "类别",
        "一句话摘要", "内容角度", "是否需人工核验", "人工核验原因", "重要性评分", "状态", "推荐平台",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "当天推送信息"
    ws.append(headers)
    for row in _dedupe_pushed_rows(selection):
        ws.append([
            _feedback_id(row),
            "",
            "",
            "",
            f"{run_date:%Y-%m-%d}",
            row.get("_excel_row", ""),
            row.get("title", ""),
            row.get("source", ""),
            clean_article_url(row.get("url", "")),
            row.get("published_date", ""),
            row.get("region", ""),
            row.get("category", ""),
            row.get("summary_cn", ""),
            row.get("content_angle", ""),
            row.get("need_verify", ""),
            _attachment_verify_reason(row),
            row.get("importance_score", ""),
            row.get("status", ""),
            row.get("platform", ""),
        ])
    ws.freeze_panes = "A2"
    like_validation = DataValidation(type="list", formula1='"重点关注,可做选题,一般参考,无需关注"', allow_blank=True)
    reason_validation = DataValidation(
        type="list",
        formula1='"中国创新药,NMPA / CDE,BD / 授权 / 出海,融资并购,商业化 / 医保 / 集采,上市公司公告,产业事件明确,信息不够新,偏研发进展,偏基础科研,无明确事件,来源质量较低,链接无法打开,重复资讯"',
        allow_blank=True,
    )
    ws.add_data_validation(like_validation)
    ws.add_data_validation(reason_validation)
    like_validation.add(f"B2:B{max(ws.max_row, 2)}")
    reason_validation.add(f"C2:C{max(ws.max_row, 2)}")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    widths = [14, 18, 20, 28, 12, 10, 42, 24, 52, 14, 10, 22, 56, 48, 16, 48, 12, 14, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    wb.save(path)
    return path


def _save_send_record(output_dir: Path, run_date, selection, mail_result, paths: dict[str, Path]) -> Path:
    """保存发送记录 CSV，方便追踪每天推送了哪些资讯。"""
    record_path = output_dir / f"send_record_{run_date:%Y%m%d}.csv"
    rows = _dedupe_pushed_rows(selection)
    with record_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["run_date", "excel_row", "title", "source", "url", "mail_sent", "mail_message", "text_path", "html_path", "pushed_excel_path"])
        for row in rows:
            writer.writerow([
                f"{run_date:%Y-%m-%d}",
                row.get("_excel_row", ""),
                row.get("title", ""),
                row.get("source", ""),
                clean_article_url(row.get("url", "")),
                "是" if mail_result.sent else "否",
                mail_result.message,
                paths.get("text", ""),
                paths.get("html", ""),
                paths.get("pushed_excel", ""),
            ])
    return record_path


def _save_source_health(output_dir: Path, run_date, config: dict, fetch_errors: list[str], fetched_count: int, appended: int) -> Path:
    """保存每日来源健康记录，用于持续淘汰失效来源并补充高质量来源。"""
    path = output_dir / f"source_health_{run_date:%Y%m%d}.csv"
    error_sources = {item.split(" 拉取失败", 1)[0]: item for item in fetch_errors}
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["run_date", "source", "enabled", "status", "message", "total_fetched", "total_appended"])
        for source in config.get("fetch", {}).get("sources", []):
            if source.get("enabled") is False:
                continue
            name = source.get("name", "")
            error = error_sources.get(name, "")
            writer.writerow([
                f"{run_date:%Y-%m-%d}", name, "是", "失败" if error else "正常",
                error, fetched_count, appended,
            ])
    return path


def main() -> int:
    args = parse_args()
    run_date = parse_run_date(args.date)
    config_path = resolve_path(args.config, BASE_DIR)
    config = load_config(config_path)

    output_dir = (BASE_DIR / config.get("output", {}).get("output_dir", "outputs")).resolve()
    log_dir = (BASE_DIR / "logs").resolve()
    backup_dir = (BASE_DIR / config.get("excel", {}).get("backup_dir", "backups")).resolve()
    ensure_dirs(output_dir, log_dir, backup_dir)
    ensure_dirs(BASE_DIR / "feedback")
    logger = setup_logging(log_dir, run_date, args.verbose)
    logger.info(sync_feedback(BASE_DIR))
    if config.get("feedback", {}).get("collect_from_mail", False):
        feedback_result = collect_feedback_mail(BASE_DIR, output_dir, config, run_date)
        logger.info(feedback_result.message)
    config["_feedback_weights"] = load_feedback_weights(BASE_DIR, output_dir)
    trusted_feedback, suspicious_feedback = audit_feedback_history(BASE_DIR)
    logger.info("反馈质量审核：可信 %s 条；可疑隔离 %s 条", len(trusted_feedback), len(suspicious_feedback))
    load_dotenv(BASE_DIR / ".env")
    config["_feedback_mail_to"] = os.getenv("FEEDBACK_MAIL_TO", os.getenv("MAIL_FROM", "")).strip()
    logger.info("已读取历史反馈偏好：来源 %s 个；类别 %s 个；原因关键词 %s 个",
                len(config["_feedback_weights"].get("source", {})),
                len(config["_feedback_weights"].get("category", {})),
                len(config["_feedback_weights"].get("keyword", {})))

    excel_path = resolve_path(args.excel, Path.cwd())
    logger.info("运行日期：%s", run_date)
    logger.info("读取 Excel 文件：%s", excel_path)

    fetch_errors: list[str] = []
    fetched_count = 0
    appended = 0
    fetch_failed_all = False
    if args.fetch or config.get("fetch", {}).get("enabled", False):
        if not excel_path.exists() and config.get("fetch", {}).get("create_excel_if_missing", True):
            logger.info("Excel 不存在，正在创建标准资讯库模板：%s", excel_path)
            ensure_news_workbook(excel_path, config.get("excel", {}).get("sheet_name", "资讯库"))
        if excel_path.exists() and config.get("excel", {}).get("backup_before_write", True):
            backup_path = backup_excel(excel_path, backup_dir)
            logger.info("自动检索写入前备份已生成：%s", backup_path)
        fetched_rows, fetch_errors, fetched_count = fetch_rss_items(config, run_date)
        for error in fetch_errors:
            logger.warning(error)
        try:
            appended, skipped_existing = append_fetched_rows(
                excel_path,
                config.get("excel", {}).get("sheet_name", "资讯库"),
                fetched_rows,
            )
        except Exception as exc:
            logger.error("自动检索结果写入 Excel 失败：%s", exc)
            return 4
        skipped_filtered = max(fetched_count - len(fetched_rows), 0)
        logger.info(
            "自动检索完成：RSS 原始条目 %s 条；过滤后 %s 条；新增写入 %s 条；URL 已存在跳过 %s 条；关键词/日期过滤跳过 %s 条。",
            fetched_count,
            len(fetched_rows),
            appended,
            skipped_existing,
            skipped_filtered,
        )
        health_path = _save_source_health(output_dir, run_date, config, fetch_errors, fetched_count, appended)
        logger.info("来源健康记录已保存：%s", health_path)
        fetch_failed_all = bool(fetch_errors and fetched_count == 0)
        if fetch_failed_all:
            logger.error("本次所有启用的检索来源均失败，将发送运行告警邮件。")
        if args.fetch_only:
            return 0

    try:
        excel_data = load_news_excel(
            excel_path,
            config.get("excel", {}).get("sheet_name", "资讯库"),
            config.get("fields", {}).get("aliases", {}),
        )
    except Exception as exc:
        logger.error("读取 Excel 失败：%s", exc)
        return 1

    logger.info("读取到资讯：%s 条", len(excel_data.rows))
    if excel_data.missing_fields:
        logger.warning("模板缺少以下可选/建议字段：%s", ", ".join(excel_data.missing_fields))
    logger.info("字段映射：%s", excel_data.reverse_map)

    if args.validate_only:
        issues = validate_rows(excel_data.rows, run_date)
        _log_issues(logger, issues)
        logger.info("validate-only 完成，共发现 %s 条提醒/问题。", len(issues))
        return 0 if not any(i.level == "ERROR" for i in issues) else 2

    selection = select_digest_items(excel_data.rows, run_date, config)
    _log_issues(logger, selection.issues)
    logger.info("今日候选资讯：%s 条", selection.total_today)
    logger.info("重点资讯：%s 条；需人工核验：%s 条；重复跳过：%s 条", len(selection.重点资讯), len(selection.需人工核验), len(selection.duplicates))
    for dup in selection.duplicates:
        logger.info("重复跳过：第 %s 行，%s", dup.get("_excel_row"), dup.get("title"))

    has_digest_items = bool(selection.重点资讯 or selection.需人工核验 or selection.待人工确认)
    if not has_digest_items and not config.get("digest", {}).get("enable_empty_digest", False):
        logger.info("没有符合推送条件的资讯；仍会生成 dry-run 日报文件，但默认不发送空日报。")

    portal_rows = _dedupe_pushed_rows(selection)
    if portal_rows and config.get("feedback", {}).get("public_portal_enabled", True):
        public_feedback_url, publish_message = publish_digest(BASE_DIR, run_date, portal_rows)
        config["_feedback_public_url"] = public_feedback_url
        logger.info(publish_message)
    if portal_rows and config.get("feedback", {}).get("local_portal_enabled", False):
        portal_path, portal_url = build_feedback_portal(
            output_dir,
            run_date,
            portal_rows,
            int(config.get("feedback", {}).get("local_portal_port", 8765)),
        )
        config["_feedback_portal_url"] = portal_url
        logger.info("批量反馈页已生成：%s", portal_path)

    text = build_plain_text(selection, run_date, config)
    html_text = build_html(selection, run_date, config)
    if fetch_failed_all:
        alert_text = "【运行告警】本次自动检索来源全部连接失败，以下日报可能没有最新资讯。请检查网络并稍后补跑。\n\n"
        text = alert_text + text
        alert_html = '<div style="max-width:900px;margin:16px auto;padding:14px 18px;background:#fff7ed;border:1px solid #fb923c;color:#9a3412;font-family:Microsoft YaHei,Arial,sans-serif;">运行告警：本次自动检索来源全部连接失败，以下日报可能没有最新资讯。请检查网络并稍后补跑。</div>'
        html_text = html_text.replace("<body>", f"<body>{alert_html}", 1)
    paths = save_digest_outputs(
        text,
        html_text,
        output_dir,
        run_date,
        config.get("output", {}).get("save_text", True),
        config.get("output", {}).get("save_html", True),
    )
    logger.info("日报已生成：%s", paths)

    attachment_paths: list[Path] = []
    if config.get("mail", {}).get("attach_pushed_excel", True):
        pushed_excel_path = _save_pushed_items_excel(output_dir, run_date, selection)
        paths["pushed_excel"] = pushed_excel_path
        attachment_paths.append(pushed_excel_path)
        logger.info("当天推送信息 Excel 附件已生成：%s", pushed_excel_path)

    subject = f"{config.get('mail', {}).get('subject_prefix', '【ZettaLab 生物医药资讯雷达日报】')}{run_date:%Y-%m-%d}"
    should_send = args.send and not args.dry_run and (
        has_digest_items
        or config.get("digest", {}).get("enable_empty_digest", False)
        or fetch_failed_all
    )
    # 命令行 --send 明确表示本次要发送，因此覆盖配置中的默认 dry_run=true。
    effective_config = dict(config)
    effective_config["mail"] = dict(config.get("mail", {}))
    if should_send:
        effective_config["mail"]["dry_run"] = False
    mail_result = send_mail(
        subject,
        text,
        html_text,
        effective_config,
        dry_run=not should_send,
        attachments=attachment_paths,
    )
    logger.info(mail_result.message)

    record_path = _save_send_record(output_dir, run_date, selection, mail_result, paths)
    logger.info("发送记录已保存：%s", record_path)

    if config.get("output", {}).get("save_daily_excel_snapshot", True) and excel_path.exists():
        prefix = config.get("output", {}).get("daily_excel_prefix", "auto_news_cn")
        snapshot_path = output_dir / f"{prefix}_{run_date:%Y%m%d}.xlsx"
        try:
            shutil.copy2(excel_path, snapshot_path)
            logger.info("每日日期命名 Excel 快照已保存：%s", snapshot_path)
        except PermissionError:
            logger.warning("每日 Excel 快照保存失败：%s 可能正被打开占用；日报已正常生成。", snapshot_path)
        except Exception as exc:
            logger.warning("每日 Excel 快照保存失败：%s；日报已正常生成。", exc)

    if mail_result.sent:
        pushed_rows = [int(row["_excel_row"]) for row in selection.重点资讯]
        if pushed_rows:
            try:
                backup_path = None
                if config.get("excel", {}).get("backup_before_write", True):
                    backup_path = backup_excel(excel_path, backup_dir)
                    logger.info("写回前备份已生成：%s", backup_path)
                write_push_status(
                    excel_data,
                    excel_path,
                    pushed_rows,
                    datetime.now(),
                    clear_push_today=config.get("excel", {}).get("clear_push_today_after_send", False),
                )
                logger.info("Excel 写回成功，已更新 %s 条重点资讯。", len(pushed_rows))
            except Exception as exc:
                logger.error("Excel 写回失败，原文件未确认更新：%s", exc)
                return 3
    else:
        logger.info("未发送邮件，因此不写回 pushed/pushed_at。")

    return 0


if __name__ == "__main__":
    sys.exit(main())
