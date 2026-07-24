from __future__ import annotations

import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ExcelData:
    workbook: Any
    sheet: Worksheet
    rows: list[dict[str, Any]]
    field_map: dict[str, str]
    reverse_map: dict[str, str]
    header_row: int
    row_numbers: list[int]
    missing_fields: list[str]


def _build_field_map(headers: list[str], aliases: dict[str, list[str]]) -> tuple[dict[str, str], dict[str, str], list[str]]:
    """根据配置中的中文/英文别名，把 Excel 表头映射到标准字段名。"""
    normalized_headers = {str(h).strip(): str(h).strip() for h in headers if h is not None and str(h).strip()}
    field_map: dict[str, str] = {}
    reverse_map: dict[str, str] = {}
    missing: list[str] = []

    for canonical, names in aliases.items():
        matched = None
        for name in names:
            if name in normalized_headers:
                matched = normalized_headers[name]
                break
        if matched:
            field_map[matched] = canonical
            reverse_map[canonical] = matched
        else:
            missing.append(canonical)
    return field_map, reverse_map, missing


def load_news_excel(excel_path: Path, sheet_name: str, aliases: dict[str, list[str]]) -> ExcelData:
    """读取资讯库工作表，保留 workbook 对象用于后续原格式写回。"""
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在：{excel_path}")

    workbook = load_workbook(excel_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表不存在：{sheet_name}；当前工作表：{', '.join(workbook.sheetnames)}")

    sheet = workbook[sheet_name]
    header_row = 1
    headers = [cell.value for cell in sheet[header_row]]
    field_map, reverse_map, missing = _build_field_map(headers, aliases)

    rows: list[dict[str, Any]] = []
    row_numbers: list[int] = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        values = {headers[col_idx]: sheet.cell(row=row_idx, column=col_idx + 1).value for col_idx in range(len(headers)) if headers[col_idx]}
        if not any(v not in (None, "") for v in values.values()):
            continue
        canonical_row = {canonical: values.get(original) for original, canonical in field_map.items()}
        canonical_row["_excel_row"] = row_idx
        rows.append(canonical_row)
        row_numbers.append(row_idx)

    return ExcelData(workbook, sheet, rows, field_map, reverse_map, header_row, row_numbers, missing)


def _column_index(sheet: Worksheet, header_name: str, header_row: int = 1) -> int | None:
    """查找某个表头所在列号。"""
    for cell in sheet[header_row]:
        if str(cell.value).strip() == header_name:
            return cell.column
    return None


def backup_excel(excel_path: Path, backup_dir: Path) -> Path:
    """写回前备份原始 Excel，避免误操作后无法恢复。"""
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{excel_path.stem}_{stamp}{excel_path.suffix}"
    shutil.copy2(excel_path, backup_path)
    return backup_path


def write_push_status(
    excel_data: ExcelData,
    excel_path: Path,
    pushed_row_numbers: list[int],
    pushed_at: datetime,
    clear_push_today: bool = False,
) -> None:
    """发送成功后只写回 pushed、pushed_at 和可选 push_today 字段，不改动其他工作表。"""
    sheet = excel_data.sheet
    pushed_header = excel_data.reverse_map.get("pushed")
    pushed_at_header = excel_data.reverse_map.get("pushed_at")
    push_today_header = excel_data.reverse_map.get("push_today")

    if not pushed_header or not pushed_at_header:
        raise ValueError("缺少 pushed 或 pushed_at 字段，无法写回推送状态。")

    pushed_col = _column_index(sheet, pushed_header, excel_data.header_row)
    pushed_at_col = _column_index(sheet, pushed_at_header, excel_data.header_row)
    push_today_col = _column_index(sheet, push_today_header, excel_data.header_row) if push_today_header else None

    for row_idx in pushed_row_numbers:
        sheet.cell(row=row_idx, column=pushed_col).value = "是"
        sheet.cell(row=row_idx, column=pushed_at_col).value = pushed_at.strftime("%Y-%m-%d %H:%M:%S")
        if clear_push_today and push_today_col:
            sheet.cell(row=row_idx, column=push_today_col).value = "否"

    excel_data.workbook.save(excel_path)
