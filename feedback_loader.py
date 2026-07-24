from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from utils import normalize_text


POSITIVE_VALUES = {"喜欢", "是", "yes", "like", "推荐"}
NEGATIVE_VALUES = {"不喜欢", "否", "no", "dislike", "不推荐"}


def audit_feedback_history(base_dir: Path) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """审计反馈可信度，隔离高频、重复或极端单一的可疑反馈。"""
    history_path = base_dir / "feedback" / "feedback_history.csv"
    if not history_path.exists():
        return [], []
    with history_path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    latest: dict[tuple[str, str], dict[str, str]] = {}
    for row in rows:
        respondent = normalize_text(row.get("respondent_id")) or normalize_text(row.get("message_id"))
        feedback_id = normalize_text(row.get("feedback_id"))
        latest[(respondent, feedback_id)] = row

    by_respondent: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in latest.values():
        respondent = normalize_text(row.get("respondent_id")) or normalize_text(row.get("message_id"))
        by_respondent[respondent].append(row)

    burst_groups: dict[tuple[str, str], list[datetime]] = defaultdict(list)
    for row in latest.values():
        try:
            timestamp = datetime.fromisoformat(normalize_text(row.get("collected_date")).replace("Z", "+00:00"))
            burst_groups[(normalize_text(row.get("feedback_id")), normalize_text(row.get("action")))].append(timestamp)
        except ValueError:
            pass
    burst_keys = {
        key for key, times in burst_groups.items()
        if len(times) >= 10 and (max(times) - min(times)).total_seconds() < 300
    }

    trusted: list[dict[str, str]] = []
    suspicious: list[dict[str, str]] = []
    for respondent, respondent_rows in by_respondent.items():
        actions = [normalize_text(row.get("action")) for row in respondent_rows]
        same_action_ratio = max((actions.count(action) for action in set(actions)), default=0) / max(len(actions), 1)
        times = []
        for row in respondent_rows:
            try:
                times.append(datetime.fromisoformat(normalize_text(row.get("collected_date")).replace("Z", "+00:00")))
            except ValueError:
                pass
        rapid_batch = len(times) >= 15 and (max(times) - min(times)).total_seconds() < 60
        excessive = len(respondent_rows) > 60
        extreme_uniform = len(respondent_rows) >= 12 and same_action_ratio >= 0.98
        for row in respondent_rows:
            audited = dict(row)
            reasons = []
            if rapid_batch:
                reasons.append("60秒内高频批量反馈")
            if excessive:
                reasons.append("单用户反馈数量异常")
            if extreme_uniform:
                reasons.append("大量反馈选项高度单一")
            if (normalize_text(row.get("feedback_id")), normalize_text(row.get("action"))) in burst_keys:
                reasons.append("同一资讯出现跨用户短时集中同向反馈")
            audited["_audit_reason"] = "；".join(reasons)
            if reasons:
                suspicious.append(audited)
            else:
                trusted.append(audited)
    return trusted, suspicious


def save_feedback_audit_report(base_dir: Path, trusted: list[dict[str, str]], suspicious: list[dict[str, str]]) -> Path:
    """保存反馈质量审计报告，便于运营人工复查异常反馈。"""
    path = base_dir / "outputs" / "feedback_audit.csv"
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = ["collected_date", "respondent_id", "feedback_id", "action", "tags", "title", "source", "category", "_audit_reason"]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in trusted:
            record = dict(row)
            record["_audit_reason"] = "可信反馈"
            writer.writerow(record)
        for row in suspicious:
            writer.writerow(row)
    return path


def load_feedback_weights(base_dir: Path, output_dir: Path) -> dict[str, dict[str, int]]:
    """读取用户填写过的推送附件，并汇总来源、类别和关键词偏好权重。"""
    files = list((base_dir / "feedback").glob("*.xlsx"))
    files.extend(output_dir.glob("pushed_items_*.xlsx"))
    source_weights: dict[str, int] = defaultdict(int)
    category_weights: dict[str, int] = defaultdict(int)
    keyword_weights: dict[str, int] = defaultdict(int)

    trusted_history, suspicious_history = audit_feedback_history(base_dir)
    save_feedback_audit_report(base_dir, trusted_history, suspicious_history)
    for row in trusted_history:
        action = normalize_text(row.get("action"))
        value_scores = {"重点关注": 3, "可做选题": 2, "一般参考": 0, "无需关注": -3}
        if action in value_scores:
            delta = value_scores[action]
        else:
            preference = normalize_text(row.get("preference")).lower()
            delta = 2 if preference in POSITIVE_VALUES else -2 if preference in NEGATIVE_VALUES else 0
        if not delta:
            continue
        source = normalize_text(row.get("source")).lower()
        category = normalize_text(row.get("category")).lower()
        try:
            tags = json.loads(normalize_text(row.get("tags")) or "[]")
        except json.JSONDecodeError:
            tags = []
        quality_tags = {"来源质量较低", "链接无法打开", "重复资讯"}
        topic_tags = [normalize_text(tag).lower() for tag in tags if normalize_text(tag) and normalize_text(tag) not in quality_tags]
        if source:
            source_delta = delta
            if "来源质量较低" in tags or "链接无法打开" in tags:
                source_delta -= 3
            if "重复资讯" in tags:
                source_delta -= 1
            source_weights[source] += source_delta
        if category and not any(tag in tags for tag in quality_tags):
            category_weights[category] += delta
        for tag in topic_tags:
            keyword_weights[tag] += delta

    for path in files:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [normalize_text(cell.value) for cell in ws[1]]
            index = {header: idx for idx, header in enumerate(headers)}
            feedback_col = index.get("是否喜欢该资讯")
            if feedback_col is None:
                continue
            for values in ws.iter_rows(min_row=2, values_only=True):
                feedback = normalize_text(values[feedback_col]).lower()
                if feedback in POSITIVE_VALUES:
                    delta = 2
                elif feedback in NEGATIVE_VALUES:
                    delta = -2
                else:
                    continue
                source = normalize_text(values[index["来源"]]) if "来源" in index else ""
                category = normalize_text(values[index["类别"]]) if "类别" in index else ""
                reason = normalize_text(values[index["反馈原因"]]) if "反馈原因" in index else ""
                if source:
                    source_weights[source.lower()] += delta
                if category:
                    category_weights[category.lower()] += delta
                for term in reason.replace("，", ",").replace("；", ",").split(","):
                    term = normalize_text(term).lower()
                    if term:
                        keyword_weights[term] += delta
        except Exception:
            continue

    return {
        "source": dict(source_weights),
        "category": dict(category_weights),
        "keyword": dict(keyword_weights),
    }


def feedback_rank_bonus(row: dict[str, Any], weights: dict[str, dict[str, int]]) -> int:
    """根据历史喜欢/不喜欢反馈，为候选资讯提供排序加减分。"""
    source = normalize_text(row.get("source")).lower()
    category = normalize_text(row.get("category")).lower()
    text = f"{normalize_text(row.get('title'))} {normalize_text(row.get('summary_cn'))} {category}".lower()
    bonus = weights.get("source", {}).get(source, 0)
    bonus += weights.get("category", {}).get(category, 0)
    bonus += sum(value for term, value in weights.get("keyword", {}).items() if term and term in text)
    return max(-10, min(10, bonus))
