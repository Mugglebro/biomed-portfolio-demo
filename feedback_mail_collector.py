from __future__ import annotations

import csv
import imaplib
import os
import poplib
import re
from dataclasses import dataclass
from datetime import date, timedelta
from email import message_from_bytes
from email.header import decode_header
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

from utils import normalize_text


SUBJECT_PATTERN = re.compile(r"^\[Radar反馈\]\s*(喜欢|不喜欢|链接问题|方向不符)\s+([A-F0-9]{8})", re.IGNORECASE)


@dataclass
class FeedbackCollectResult:
    collected: int
    skipped: int
    message: str


def _decode_subject(value: str | None) -> str:
    chunks = []
    for part, encoding in decode_header(value or ""):
        if isinstance(part, bytes):
            chunks.append(part.decode(encoding or "utf-8", errors="replace"))
        else:
            chunks.append(part)
    return "".join(chunks).strip()


def _load_feedback_index(output_dir: Path) -> dict[str, dict[str, str]]:
    """从已发送附件中建立反馈编号到资讯信息的映射。"""
    index: dict[str, dict[str, str]] = {}
    for path in sorted(output_dir.glob("pushed_items_*.xlsx"), reverse=True):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [normalize_text(cell.value) for cell in ws[1]]
            cols = {header: idx for idx, header in enumerate(headers)}
            if "反馈编号" not in cols:
                continue
            for values in ws.iter_rows(min_row=2, values_only=True):
                feedback_id = normalize_text(values[cols["反馈编号"]]).upper()
                if not feedback_id or feedback_id in index:
                    continue
                index[feedback_id] = {
                    "title": normalize_text(values[cols["标题"]]) if "标题" in cols else "",
                    "source": normalize_text(values[cols["来源"]]) if "来源" in cols else "",
                    "category": normalize_text(values[cols["类别"]]) if "类别" in cols else "",
                    "url": normalize_text(values[cols["原文链接"]]) if "原文链接" in cols else "",
                }
        except Exception:
            continue
    return index


def _feedback_row_from_message(raw_message: bytes, fallback_id: str, feedback_index: dict[str, dict[str, str]], run_date: date) -> tuple[str, list[str]] | None:
    """解析一封邮件；只有符合反馈主题格式的邮件才会返回记录。"""
    msg = message_from_bytes(raw_message)
    message_id = normalize_text(msg.get("Message-ID")) or fallback_id
    subject = _decode_subject(msg.get("Subject"))
    match = SUBJECT_PATTERN.match(subject)
    if not match:
        return None
    action = match.group(1)
    feedback_id = match.group(2).upper()
    info = feedback_index.get(feedback_id, {})
    preference = "喜欢" if action == "喜欢" else "不喜欢"
    return message_id, [
        f"{run_date:%Y-%m-%d}", message_id, feedback_id, action, preference,
        info.get("title", ""), info.get("source", ""), info.get("category", ""), info.get("url", ""),
    ]


def collect_feedback_mail(base_dir: Path, output_dir: Path, config: dict, run_date: date) -> FeedbackCollectResult:
    """通过 IMAP 自动读取反馈邮件，并保存为本地 CSV 历史。"""
    feedback_cfg = config.get("feedback", {})
    if not feedback_cfg.get("enabled", True) or not feedback_cfg.get("collect_from_mail", True):
        return FeedbackCollectResult(0, 0, "反馈邮件采集未启用。")

    load_dotenv(base_dir / ".env")
    host = os.getenv("IMAP_HOST", "imap.163.com").strip()
    port = int(os.getenv("IMAP_PORT", "993") or "993")
    user = os.getenv("IMAP_USER", os.getenv("SMTP_USER", "")).strip()
    password = os.getenv("IMAP_PASSWORD", os.getenv("SMTP_PASSWORD", "")).strip()
    if not host or not user or not password:
        return FeedbackCollectResult(0, 0, "IMAP 配置不完整，已跳过反馈邮件采集。")

    history_path = base_dir / "feedback" / "feedback_history.csv"
    history_path.parent.mkdir(parents=True, exist_ok=True)
    processed_ids: set[str] = set()
    if history_path.exists():
        with history_path.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                processed_ids.add(normalize_text(row.get("message_id")))

    feedback_index = _load_feedback_index(output_dir)
    collected = 0
    skipped = 0
    new_rows: list[list[str]] = []
    protocol = "IMAP"
    try:
        with imaplib.IMAP4_SSL(host, port) as client:
            client.login(user, password)
            client.select("INBOX", readonly=True)
            since = (run_date - timedelta(days=int(feedback_cfg.get("mail_scan_days", 30)))).strftime("%d-%b-%Y")
            status, data = client.search(None, "SINCE", since)
            if status != "OK":
                return FeedbackCollectResult(0, 0, "IMAP 搜索反馈邮件失败。")
            for number in data[0].split():
                status, payload = client.fetch(number, "(RFC822)")
                if status != "OK" or not payload or not isinstance(payload[0], tuple):
                    continue
                parsed = _feedback_row_from_message(payload[0][1], number.decode(), feedback_index, run_date)
                if not parsed:
                    continue
                message_id, history_row = parsed
                if message_id in processed_ids:
                    skipped += 1
                    continue
                new_rows.append(history_row)
                processed_ids.add(message_id)
                collected += 1
    except Exception as imap_exc:
        protocol = "POP3"
        pop_host = os.getenv("POP3_HOST", "pop.163.com").strip()
        pop_port = int(os.getenv("POP3_PORT", "995") or "995")
        try:
            client = poplib.POP3_SSL(pop_host, pop_port, timeout=20)
            try:
                client.user(user)
                client.pass_(password)
                count, _ = client.stat()
                max_messages = int(feedback_cfg.get("pop3_scan_messages", 200))
                for number in range(max(1, count - max_messages + 1), count + 1):
                    _, lines, _ = client.retr(number)
                    parsed = _feedback_row_from_message(b"\r\n".join(lines), str(number), feedback_index, run_date)
                    if not parsed:
                        continue
                    message_id, history_row = parsed
                    if message_id in processed_ids:
                        skipped += 1
                        continue
                    new_rows.append(history_row)
                    processed_ids.add(message_id)
                    collected += 1
            finally:
                try:
                    client.quit()
                except Exception:
                    pass
        except Exception as pop_exc:
            return FeedbackCollectResult(0, 0, f"反馈邮件采集失败：IMAP={imap_exc}；POP3={pop_exc}")

    file_exists = history_path.exists()
    with history_path.open("a", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["collected_date", "message_id", "feedback_id", "action", "preference", "title", "source", "category", "url"])
        writer.writerows(new_rows)
    return FeedbackCollectResult(collected, skipped, f"反馈邮件采集完成（{protocol}）：新增 {collected} 条，已处理跳过 {skipped} 条。")
