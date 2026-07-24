from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).resolve().parents[1]
OUT = BASE_DIR / "sample_news.xlsx"


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "资讯库"
    headers = [
        "编号", "标题", "来源", "原文链接", "发布时间", "收集时间", "语言", "地区", "类别",
        "中文摘要", "关键实体", "与ZettaLab相关性", "重要性评分", "内容角度", "推荐平台",
        "状态", "是否需人工核验", "是否今日推送", "是否已推送", "推送时间", "人工备注",
    ]
    ws.append(headers)
    today = date.today().strftime("%Y-%m-%d")
    rows = [
        [1, "FDA 发布某类 AI 医疗软件监管更新", "FDA", "https://example.com/fda-ai-medical-software", today, today, "英文", "美国", "监管", "FDA 对 AI 医疗软件监管文件进行更新，需结合原文确认适用范围。", "FDA; AI medical software", "中", 4, "AI 医疗软件监管边界", "LinkedIn", "建议推送", "是", "是", "否", "", "需核验监管原文"],
        [2, "大型药企宣布 AI 制药研发合作", "公司新闻稿", "https://example.com/ai-drug-discovery-partnership", today, today, "英文", "全球", "AI 制药", "合作显示 AI 工具正在向药物发现流程中的具体环节渗透。", "PharmaCo; AI Lab", "高", 5, "AI in Biopharma 从工具走向流程协同", "LinkedIn; 企业微信", "可发帖", "否", "是", "否", "", ""],
        [3, "CRISPR 基因编辑疗法更新长期随访数据", "期刊新闻", "https://example.com/crispr-follow-up", today, today, "英文", "全球", "基因编辑", "长期随访数据提示疗效和安全性仍需持续观察。", "CRISPR; gene editing", "中", 4, "基因编辑从突破转向长期证据积累", "企业微信", "建议推送", "是", "是", "否", "", "需核验临床数据口径"],
        [4, "蛋白结构预测工具发布分子生物学新功能", "研究机构", "https://example.com/protein-structure-tool", today, today, "英文", "全球", "分子生物学", "新功能聚焦蛋白结构和实验设计流程的连接。", "protein structure", "高", 4, "分子生物学工具链的流程化趋势", "LinkedIn", "建议推送", "否", "是", "否", "", ""],
        [5, "某生物技术公司完成新一轮融资并推进临床项目", "行业媒体", "https://example.com/biotech-financing-clinical", today, today, "中文", "中国", "融资", "融资金额和临床阶段信息需以官方披露核验。", "BiotechCo", "低", 4, "融资与临床推进节奏", "企业微信", "建议推送", "是", "是", "否", "", "核验融资金额、投资方和临床阶段"],
        [6, "大型药企宣布 AI 制药研发合作", "公司新闻稿", "https://example.com/ai-drug-discovery-partnership", today, today, "英文", "全球", "AI 制药", "重复 URL 测试。", "PharmaCo; AI Lab", "高", 4, "重复 URL 测试", "LinkedIn", "建议推送", "否", "是", "否", "", ""],
        [7, "实验数据管理平台更新功能", "产品博客", "", today, today, "中文", "中国", "实验数据管理", "", "ELN; LIMS", "高", 4, "实验数据管理与 AI 科研工作流", "LinkedIn", "建议推送", "否", "是", "否", "", ""],
    ]
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for column_cells in ws.columns:
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(len(str(column_cells[0].value)) + 8, 14), 36)

    for name in ["来源清单", "分类字典", "每日简报", "使用SOP"]:
        sheet = wb.create_sheet(name)
        sheet["A1"] = name
        sheet["A1"].font = Font(bold=True)

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
