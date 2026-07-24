from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).resolve().parents[1]
OUT = BASE_DIR / "real_news.xlsx"


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "资讯库"
    headers = [
        "编号", "标题", "来源", "原文链接", "发布时间", "收集时间", "语言", "地区", "类别",
        "中文摘要", "关键实体", "与ZettaLab相关性", "重要性评分", "内容角度", "推荐平台",
        "状态", "是否需人工核验", "是否今日推送", "是否已推送", "推送时间", "人工备注",
    ]
    ws.append(headers)
    collected = date.today().strftime("%Y-%m-%d")
    rows = [
        [
            1,
            "FDA 批准首个慢性丁型肝炎病毒（HDV）感染治疗药物 Hepcludex",
            "FDA",
            "https://www.fda.gov/news-events/press-announcements/fda-approves-first-treatment-chronic-hepatitis-delta-virus-hdv-infection",
            "2026-05-22",
            collected,
            "英文",
            "美国",
            "监管审批",
            "FDA 批准 Gilead 的 Hepcludex 用于成人慢性 HDV 感染，这是美国首个获批的慢性 HDV 治疗药物；疗效和适用人群需以 FDA 原文和标签为准。",
            "FDA; Hepcludex; bulevirtide; Gilead; HDV",
            "中",
            5,
            "罕见/严重感染病种的加速审批与真实临床需求",
            "企业微信",
            "建议推送",
            "是",
            "是",
            "否",
            "",
            "需核验适应症、加速审批路径、关键临床终点和安全性警示。",
        ],
        [
            2,
            "FDA 发布 Elsa 4.0 并整合 HALO 数据平台，扩大内部 AI 能力",
            "FDA",
            "https://www.fda.gov/news-events/press-announcements/fda-expands-ai-capabilities-and-completes-data-platform-consolidation",
            "2026-05-06",
            collected,
            "英文",
            "美国",
            "AI/监管科技",
            "FDA 宣布升级内部 AI 工具 Elsa 4.0，并将多个申请和提交数据源整合进 HALO 平台，体现监管机构正在把 AI 嵌入审评和运营流程。",
            "FDA; Elsa 4.0; HALO; AI",
            "高",
            5,
            "AI 从单点工具进入监管与研发数据工作流",
            "LinkedIn; 企业微信",
            "可发帖",
            "是",
            "是",
            "否",
            "",
            "需核验 FDA 对数据安全、人审参与和联网能力的原文描述。",
        ],
        [
            3,
            "EMA CHMP 2026 年 5 月会议建议 8 个新药获批，并支持口服 Wegovy 适应症扩展",
            "EMA",
            "https://www.ema.europa.eu/en/news/meeting-highlights-committee-medicinal-products-human-use-chmp-18-21-may-2026",
            "2026-05-23",
            collected,
            "英文",
            "欧洲",
            "监管审批",
            "EMA CHMP 5 月会议给出多项积极意见，包括 Etcamah、Jascayd、Vijoice 等新药建议，以及 Wegovy 口服片剂用于体重管理的适应症扩展建议。",
            "EMA; CHMP; Etcamah; Jascayd; Vijoice; Wegovy",
            "中",
            4,
            "欧洲监管审评节奏与 GLP-1 口服化趋势",
            "企业微信",
            "建议推送",
            "是",
            "是",
            "否",
            "",
            "需核验 CHMP 意见与欧盟委员会正式批准之间的差异。",
        ],
        [
            4,
            "Lantern Pharma 推出面向罕见癌症药物发现与开发的多智能体 AI co-scientist 平台 withZeta.ai",
            "Business Wire / Lantern Pharma",
            "https://www.businesswire.com/news/home/20260414869501/en/Lantern-Pharma-Launches-withZeta.ai-the-Worlds-First-Multi-Agentic-A.I.-Co-Scientist-for-Rare-Cancer-Drug-Discovery-Development-Subscriptions-Now-Open-Debut-Events-at-Nasdaq-MarketSite-on-416-and-at-AACR-2026-from-417-to-4",
            "2026-04-14",
            collected,
            "英文",
            "美国",
            "AI 制药",
            "Lantern Pharma 宣布推出 withZeta.ai，定位为用于罕见癌症药物发现与开发的多智能体 AI co-scientist；商业化效果和研发价值仍需后续数据验证。",
            "Lantern Pharma; withZeta.ai; RADR; AI co-scientist",
            "高",
            4,
            "AI in Biopharma 从模型能力转向研发流程协同产品",
            "LinkedIn",
            "建议推送",
            "是",
            "是",
            "否",
            "",
            "需核验公司自称“world's first”等表述，日报中避免直接采用宣传性措辞。",
        ],
        [
            5,
            "Cerevance 完成帕金森病 pivotal Phase 3 ARISE 试验入组，并完成 2000 万美元 C 轮融资",
            "GlobeNewswire / Cerevance",
            "https://www.globenewswire.com/news-release/2026/05/12/3292753/0/en/cerevance-completes-enrollment-in-pivotal-phase-3-parkinson-s-disease-trial-and-closes-oversubscribed-20-million-series-c-to-extend-runway-into-2027.html",
            "2026-05-12",
            collected,
            "英文",
            "美国/欧洲",
            "临床/融资",
            "Cerevance 宣布 solengepras 的帕金森病 pivotal Phase 3 ARISE 试验完成入组，并完成 2000 万美元 C 轮融资，用于支持读出前运营。",
            "Cerevance; solengepras; Parkinson's disease; Phase 3; Series C",
            "中",
            4,
            "神经退行性疾病临床资产融资与关键读出窗口",
            "企业微信",
            "建议推送",
            "是",
            "是",
            "否",
            "",
            "需核验入组人数、读出时间、融资金额和投资方。",
        ],
        [
            6,
            "Scarlet Therapeutics 宣布实验室培养通用红细胞在体内循环半衰期接近捐献血液，并完成 320 万英镑种子轮融资",
            "GlobeNewswire / Scarlet Therapeutics",
            "https://www.globenewswire.com/news-release/2026/05/07/3289639/0/en/scarlet-therapeutics-demonstrates-in-vivo-survival-of-lab-grown-universal-red-blood-cells-equivalent-to-donated-blood.html",
            "2026-05-07",
            collected,
            "英文",
            "英国",
            "细胞治疗/融资",
            "Scarlet Therapeutics 披露其实验室培养通用红细胞在体内成熟并循环，半衰期接近捐献红细胞，同时完成 320 万英镑种子轮融资推进平台开发。",
            "Scarlet Therapeutics; lab-grown RBC; universal red blood cells",
            "中",
            4,
            "工程化细胞平台从输血替代延伸到长效递送载体",
            "企业微信",
            "建议推送",
            "是",
            "是",
            "否",
            "",
            "需核验该结果属于临床前阶段，避免表述为已临床验证。",
        ],
        [
            7,
            "Northwestern 报道新实验方法可大规模分析蛋白结构能量景观",
            "Northwestern Medicine News Center",
            "https://news.feinberg.northwestern.edu/2026/05/21/new-experimental-method-reveals-protein-energy-landscapes-on-larger-scale/",
            "2026-05-21",
            collected,
            "英文",
            "美国",
            "蛋白结构/分子生物学",
            "Northwestern Medicine 报道一项用于大规模分析蛋白结构构象波动和能量景观的实验方法，可能为蛋白工程和数据驱动建模提供新数据来源。",
            "protein energy landscapes; conformational fluctuations; protein engineering",
            "高",
            4,
            "实验数据规模化对蛋白设计和 AI 结构模型的支撑价值",
            "LinkedIn; 企业微信",
            "建议推送",
            "否",
            "是",
            "否",
            "",
            "",
        ],
    ]

    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    widths = {
        "A": 8, "B": 44, "C": 24, "D": 48, "E": 14, "F": 14, "G": 10, "H": 12,
        "I": 16, "J": 52, "K": 34, "L": 16, "M": 12, "N": 34, "O": 18,
        "P": 12, "Q": 16, "R": 16, "S": 16, "T": 20, "U": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for name in ["来源清单", "分类字典", "每日简报", "使用SOP"]:
        sheet = wb.create_sheet(name)
        sheet["A1"] = name
        sheet["A1"].font = Font(bold=True)

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
